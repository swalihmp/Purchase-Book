from django.shortcuts import render,redirect,HttpResponse
from .models import products,user_data,short,Product,exp
from django.contrib import messages
from django.db.models import Q
from datetime import datetime
from django.db.models import OuterRef, Subquery
from django.http import HttpResponse,JsonResponse
import json
from openpyxl import Workbook

# Create your views here.

def home(request):
    if 'username' in request.session:
        # partyname_subquery = prod_data.objects.filter(id=OuterRef('id')).values('party')[:1]
        
        # short_queryset = short.objects.filter(status='pending').order_by('-is_emergancy').annotate(party=Subquery(partyname_subquery))

        datas = short.objects.filter(status='pending').order_by('-is_emergancy')
         
        context = {
            'datas' : datas,
            'nums': short.objects.filter(status='pending').count()
        }
        return render(request, 'home.html',context)
    else:
        return render(request,'login.html')
    
def showorder(request):
    if 'username' in request.session:
        dates = datetime.now().date()
        context = {
            'datas' : short.objects.filter(ordate__range = [dates , dates], status='orderd')
        }
        return render(request, 'showorder.html',context)
    else:
        return render(request,'login.html')
    
def login_p(request):
    if 'username' in request.session:
        context = {
            'datas' : short.objects.filter(status='pending').order_by('-is_emergancy'),
            'nums': short.objects.filter(status='pending').count()
        }
        return render(request, 'home.html',context)
    else:
        return render(request, 'login.html')
    

def login_form(request):
    if 'username' in request.session:
        context = {
            'datas' : short.objects.filter(status='pending'),
            'nums': short.objects.filter(status='pending').count()
        }
        return render(request, 'home.html',context)
    elif request.method == 'POST':
        if user_data.objects.filter(
            username = request.POST['username'], password = request.POST['password']
        ).exists():
            request.session['username'] = request.POST['username']
            context = {
                'datas' : short.objects.filter(status='pending').order_by('-is_emergancy'),
                'nums': short.objects.filter(status='pending').count()
            }
            return render(request, 'home.html',context)
            
        else:
            messages.info(request, "Invalid Credentials...!")
            return render(request,'login.html')
    else:
        return render(request,'login.html')
    
    
    
def add_prod(request):
    if 'username' in request.session:
        if request.method == 'POST':
            pname = request.POST['pname']
            
            party1 = request.POST['party1']
            party2 = request.POST['party2']
            party3 = request.POST['party3']
            party4 = request.POST['party4']
            
            p1_impo = request.POST['p1_imp']
            if p1_impo == "imp":
                p1_imp = True
            else :
                p1_imp = False
            
            p2_impo = request.POST['p2_imp']
            if p2_impo == "imp":
                p2_imp = True
            else :
                p2_imp = False
                
            p3_impo = request.POST['p3_imp']
            if p3_impo == "imp":
                p3_imp = True
            else :
                p3_imp = False
                
            p4_impo = request.POST['p4_imp']
            if p4_impo == "imp":
                p4_imp = True
            else :
                p4_imp = False
                
            data = products.objects.filter(pname=pname)
            if data.exists():
                data.update(
                    party1 = party1,
                    party2 = party2,
                    party3 = party3,
                    party4 = party4,
                    p1_imp = p1_imp,
                    p2_imp = p2_imp,
                    p3_imp = p3_imp,
                    p4_imp = p4_imp
                )
            else:
                products.objects.create(
                    pname = pname,
                    party1 = party1,
                    party2 = party2,
                    party3 = party3,
                    party4 = party4,
                    p1_imp = p1_imp,
                    p2_imp = p2_imp,
                    p3_imp = p3_imp,
                    p4_imp = p4_imp
                )

            context={
                'datas' : products.objects.all()
            }
            return render(request, 'add_prod.html',context)
        else:
            context={
                'datas' : products.objects.all()
            }
            return render(request, 'add_prod.html',context)
    else:
        return render(request, 'login.html')
    
    
def load_data(request):
    body = json.loads(request.body)
        
    pname = body['pname']

    datas = products.objects.get(pname=pname)
    
    party1 = datas.party1
    party2 = datas.party2
    party3 = datas.party3
    party4 = datas.party4
    p1_imp = datas.p1_imp
    p2_imp = datas.p2_imp
    p3_imp = datas.p3_imp
    p4_imp = datas.p4_imp
    
    
    data ={
        'party1' : party1,
        'party2' : party2,
        'party3' : party3,
        'party4' : party4,
        'p1_imp' : p1_imp,
        'p2_imp' : p2_imp,
        'p3_imp' : p3_imp,
        'p4_imp' : p4_imp,
    }
    return JsonResponse(data)
    
def add_short(request):
    if 'username' in request.session:
        context = {
            'users': user_data.objects.all(),
            'datas' : products.objects.all().values('pname').distinct(),
        }
        return render(request, 'add_short.html',context)
    else:
        return render(request,'login.html')
        
        
def add_exp(request):
    if request.method == 'POST':
        product = request.POST['pname']
        pname = products.objects.get(pname=product)
        batch = request.POST['batch']
        qty = request.POST['qty']
        date = request.POST['date']
        values = exp(
            pname = pname, 
            batch = batch,
            qty = qty,
            date = date,
        )
        values.save()
        
        context = {
                'users': user_data.objects.all(),
                'datas' : products.objects.all().values('pname').distinct(),
            }
        return render(request, 'add_exp.html',context)
    else:
        if 'username' in request.session:
            context = {
                'users': user_data.objects.all(),
                'datas' : products.objects.all().values('pname').distinct(),
            }
            return render(request, 'add_exp.html',context)
        else:
            return render(request,'login.html')
    

        
def logout_p(request):
    if 'username' in request.session:
        request.session.flush()
        return redirect('login_p')


def addshort_form(request):
    if request.method == 'POST':
        
        product = request.POST['pname']
        pname = products.objects.get(pname=product)
        
        
        if short.objects.filter(
            pname = pname, status = 'pending'
        ).exists():
            
            messages.error(request, "Medicine Already in Short List....!")
            context = {
                'users': user_data.objects.all(),
                'datas' : products.objects.all().values('pname').distinct(),
            }
            return render(request, 'add_short.html',context)
        
        else:
            is_emergancy = request.POST['emerg']
            product = request.POST['pname']
            pname = products.objects.get(pname=product)
            
            if is_emergancy == "urgent":
                values = short(
                    pname = pname, 
                    qty = request.POST['qty'],
                    is_emergancy = True,
                )
                values.save()
                return redirect('add_short')
            else:
                values = short(
                    pname = pname, 
                    qty = request.POST['qty'],
                )
                values.save()
                return redirect('add_short') 
    context = {
        'datas' : short.objects.filter(status='pending').order_by('-is_emergancy'),
    }
    return render(request,'home.html',context)

def update_sh(request,id):
    if 'username' in request.session:
        state = "orderd"
        odate = datetime.now().date()
        ex1 = short.objects.filter(id=id).update(
                status = state,
                ordate = odate
            )
        context = {
            'datas' : short.objects.filter(status='pending').order_by('-is_emergancy'),
            'nums': short.objects.filter(status='pending').count()
        }
        return render(request, 'home.html',context)
    else:
        context = {
            'datas' : short.objects.filter(status='pending').order_by('-is_emergancy'),
            'nums': short.objects.filter(status='pending').count()
        }
        return render(request, 'home.html',context)
    
def update_sh1(request,id):
    if 'username' in request.session:
        state = "pending"
        ex1 = short.objects.filter(id=id).update(
                status = state
            )
        
        dates = datetime.now().date()
        context = {
            'datas' : short.objects.filter(ordate__range = [dates , dates], status='orderd')
        }
        return render(request, 'showorder.html',context)
    else:
        return render(request, 'home.html',context)
    
def searby_item(request):
    if 'username' in request.session:
        if request.method == 'POST':
            searched = request.POST['searched']
            multiple_query = Q(Q(pname__icontains=searched) & Q(status = 'orderd'))
            datas = short.objects.filter(multiple_query)
        
            context = {
                'datas': datas
            }
        return render(request, 'showorder.html',context)
    else:
        return render(request,'login.html')
        
        
def exp_list(request):
    datas = exp.objects.filter(status='pending')
    context = {
        'datas': datas
    }
    
    return render(request, 'exp_list.html',context)
    
def update_exp(request,id):
    status = "dispatched"
    ex1 = exp.objects.filter(id=id).update(
            status = status
        )
    context = {
        'datas' : exp.objects.filter(status='pending')
    }
    return redirect('exp_list')
    
def add_party(request):
    context = {
        'datas' : exp.objects.all().filter(party='None',status='pending'),
    }
    return render(request, 'add_party.html',context)
    

def add_party_form(request):
    if request.method == 'POST':
        
        product = request.POST['pname']
        party = request.POST['party']
        pname = products.objects.get(pname=product)
        
        ex1 = exp.objects.filter(pname=pname).update(
            party = party
        )
        return redirect('exp_list')
    
def export_to_excel(request):
    # Fetch the data from your table
    table_data = exp.objects.filter(status='pending')

    # Create a new workbook and get the active worksheet
    workbook = Workbook()
    worksheet = workbook.active

    # Write the table data to the worksheet
    for row_num, data in enumerate(table_data, start=1):
        worksheet.cell(row=row_num, column=1, value=data.pname.pname)
        worksheet.cell(row=row_num, column=2, value=data.qty)
        worksheet.cell(row=row_num, column=3, value=data.date)
        worksheet.cell(row=row_num, column=4, value=data.party)
        
        # Repeat the above line for each field in your table

    # Create the response object with the appropriate content type
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=table_data.xlsx'

    # Save the workbook to the response object
    workbook.save(response)

    return response
    
def searby_date(request):
    if 'username' in request.session:
        if request.method == 'POST':
            sdate = request.POST['sdate']
            edate = request.POST['edate']
            datas= short.objects.filter(status = 'orderd', ordate__range = [sdate , edate])
            
            context = {
                'datas': datas
            }
            
        return render(request, 'showorder.html',context)
    else:
        return render(request,'login.html')
    
    